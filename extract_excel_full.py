import openpyxl

wb_f = openpyxl.load_workbook('01.  Control de producción Envasado Motupe 2026.xlsx', data_only=False)
wb_v = openpyxl.load_workbook('01.  Control de producción Envasado Motupe 2026.xlsx', data_only=True)

print('SHEET NAMES:', wb_f.sheetnames)

for sheet_name in wb_f.sheetnames:
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]
    print(f'\n{"="*60}')
    print(f'SHEET: {sheet_name}')
    print(f'Max row: {ws_f.max_row}, Max col: {ws_f.max_column}')
    
    max_r = min(ws_f.max_row, 50)
    max_c = min(ws_f.max_column, 30)
    
    print(f'\n--- FORMULAS ---')
    for r in range(1, max_r + 1):
        row_items = []
        for c in range(1, max_c + 1):
            val = ws_f.cell(r, c).value
            if val is not None:
                col_letter = openpyxl.utils.get_column_letter(c)
                row_items.append(f'{col_letter}{r}={val}')
        if row_items:
            print(f'  {" | ".join(row_items)}')
    
    print(f'\n--- VALUES ---')
    for r in range(1, max_r + 1):
        row_items = []
        for c in range(1, max_c + 1):
            val = ws_v.cell(r, c).value
            if val is not None:
                col_letter = openpyxl.utils.get_column_letter(c)
                row_items.append(f'{col_letter}{r}={val}')
        if row_items:
            print(f'  {" | ".join(row_items)}')
