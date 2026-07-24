import openpyxl
wb_f = openpyxl.load_workbook('01.  Control de producción Envasado Motupe 2026.xlsx', data_only=False)
wb_v = openpyxl.load_workbook('01.  Control de producción Envasado Motupe 2026.xlsx', data_only=True)

print('========== FLUJO DE DATOS EN EL EXCEL ==========')

ws3f = wb_f['HORA HORA L3']
ws3v = wb_v['HORA HORA L3']

for c in range(6, 12):
    hora = ws3f.cell(23, c).value
    gly_f = ws3f.cell(25, c).value
    gly_v = ws3v.cell(25, c).value
    bot_f = ws3f.cell(24, c).value
    bot_v = ws3v.cell(24, c).value
    
    print(f'Hora {hora}:')
    print(f'  GLY(r25) formula: {gly_f} | valor: {gly_v}')
    print(f'  Botellas(r24) formula: {bot_f} | valor: {bot_v}')
